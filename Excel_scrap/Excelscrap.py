from openpyxl import load_workbook
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain.docstore.document import Document
class Excelscrap:
    def Process_Excel(excel_path):
        # 載入 Excel 工作簿
        excel = load_workbook(filename=excel_path, read_only=True)
        # 初始化變數以存儲提取的文字
        text_content = ""
        # 遍歷每一頁並提取文字
        for sheet in excel:
            for row in sheet.iter_rows(values_only=True):
                text_content += " ".join(str(cell) for cell in row if cell is not None) + "\n"
        docs = Document(page_content=text_content)
        text_splitter = RecursiveCharacterTextSplitter(chunk_size=250, chunk_overlap=24)
        documents = text_splitter.split_documents([docs])
        return documents

            